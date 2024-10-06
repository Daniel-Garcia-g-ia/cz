import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill

def file_turn_report_excel(path:str, new_path:str, turn:str, day:str, data: str ):
    '''
    Opens an Excel file, creates a new sheet named based on the provided day and turno (shift), 
    copies the contents and formatting from an existing sheet, and writes new data into the new sheet.

    Args:
        path (str): The file path of the Excel workbook to open.
        new_path (str): The path where the modified Excel workbook will be saved.
        turn (str): The turno (shift) identifier to name the new sheet.
        day (str): The day identifier to name the new sheet.
        data (list): A list of data to write into the new sheet. Each element represents a row, and 
                     each row is a list of cell values.

    Returns:
        None: The function modifies the Excel file and saves it to the specified location.

    Raises:
        FileNotFoundError: If the Excel file is not found at the specified path.
        Exception: If an error occurs while opening, modifying, or saving the Excel file.
    
    Description:
        - The function opens the Excel workbook located at `path`.
        - It accesses the 'Format' sheet and creates a new sheet with the name of the provided in `turn, day`.
        - The function copies the content and style from the 'Format' sheet to the new sheet.
        - It writes the provided `data` into the new sheet starting at row 14.
        - The function merges specific cells in the new sheet as per the predefined structure.
        - Finally, it saves the modified workbook at `new_path` with a filename containing the month and closes the workbook.

    Example:
        file_turn_report_excel(
            path='reports/inspection_report.xlsx',
            new_path='reports/modified_reports',
            turn='1',
            day='2024-10-05',
            data=[['Item1', 'Value1'], ['Item2', 'Value2']],
        )

    The above example opens the file 'inspection_report.xlsx', creates a new sheet named '2024-10-05_Turn_1',
    copies the content and formatting from the 'Format' sheet, writes the provided data into the new sheet, 
    and saves the modified file in the 'modified_reports' directory.
    '''
    # define rute
    
    path_file = path
    name_new_sheet = f"{day}_Shift_{turn}"   
    
    
    try:
        #open the file
        book = openpyxl.load_workbook(path_file)
        print("Archivo abierto exitosamente.")
        #found pag day report
        sheet = book['Format']    
        
        new_sheet = book.create_sheet(title=name_new_sheet)
        
        # Copy the content and formats from the original sheet to the new sheet
        for row in sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                
                # Copying cell styles safely
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    
        
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        
        # Write data into the new sheet
        for i, fila in enumerate(data):
            for j, valor in enumerate(fila):
                new_sheet.cell(row=i+14, column=j+1, value=valor)
                
            if fila[15]>0 and fila[15]!=2:
                for j in range(1, len(fila) + 1):  
                    new_sheet.cell(row=i+14, column=j).fill = yellow_fill
                    
            if fila[15]==2:
                for j in range(1, len(fila) + 1):  
                    new_sheet.cell(row=i+14, column=j).fill = red_fill 
                
        new_sheet.merge_cells('AJ11:AQ11')
        new_sheet.merge_cells('AR11:AY11')
        new_sheet.merge_cells('BC11:BJ11')
        new_sheet.merge_cells('BL11:BS11')          
                     
        
        print("Datos escritos en la hoja con éxito.")
       
        
        book.save(f'{new_path}/Summary_inspection_line_Trad_Repor_Shift{turn}.xlsx')
        print("Cambios guardados exitosamente.") 
        
        
        book.close()
    except FileNotFoundError:
        print("El archivo no fue encontrado en la ruta especificada.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")       



def file_day_summary(path: str, new_path: str, data: list, name_new_sheet: str, month: str):
    '''
    This function opens an Excel file, creates a new sheet based on an existing "Format" sheet, copies its content and style, 
    writes data from a list into the new sheet, and saves the workbook to a new path with a specified name.

    Args:
        path (str): The path of the original Excel file to open.
        new_path (str): The path where the modified Excel file will be saved.
        data (list): A list of data (e.g., a result from an SQL query) to write into the new sheet starting at row 14.
        name_new_sheet (str): The name for the new sheet, typically representing a specific day.
        month (str): The name of the month, used to name the new saved Excel file.

    Returns:
        None: The function does not return anything, but it saves the modified Excel file in the `new_path`.

    Raises:
        FileNotFoundError: If the file at `path` is not found.
        Exception: If there is an error during file manipulation or data writing.

    Description:
        - The function opens the Excel workbook located at `path`.
        - It accesses the 'Format' sheet and creates a new sheet with the name provided in `name_new_sheet`.
        - The function copies the content and style from the 'Format' sheet to the new sheet.
        - It writes the provided `data` into the new sheet starting at row 14.
        - The function merges specific cells in the new sheet as per the predefined structure.
        - Finally, it saves the modified workbook at `new_path` with a filename containing the month and closes the workbook.

    Example:
        file_day_summary(
            path='/path/to/original.xlsx',
            new_path='/path/to/save/directory',
            data=[['Item1', 'Value1'], ['Item2', 'Value2']],
            name_new_sheet='25',
            month='October'
        )
        
    
    '''   
    
    day = name_new_sheet
    new_name_sheet = f'{day}'
    
    try:
        # Open the file
        book = openpyxl.load_workbook(path)
        print("Archivo abierto exitosamente.")
        
        # Find the format page
        sheet = book['Format']
        
        # Validate existing sheet
        if new_name_sheet in book.sheetnames:        
            del book[new_name_sheet]
        
        # Create a new sheet       
        new_sheet = book.create_sheet(title=new_name_sheet)
        
        # Copy the content and formats from the original sheet to the new sheet
        for row in sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                
                # Copying cell styles safely
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
        
        # Check if the length of the data at position 13 is equal to 33
        
        
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        
        # Write data into the new sheet
        for i, fila in enumerate(data):
            for j, valor in enumerate(fila):
                new_sheet.cell(row=i+14, column=j+1, value=valor)
                
            if fila[15]>0 and fila[15]!=2:
                for j in range(1, len(fila) + 1):  
                    new_sheet.cell(row=i+14, column=j).fill = yellow_fill
                    
            if fila[15]==2:
                for j in range(1, len(fila) + 1):  
                    new_sheet.cell(row=i+14, column=j).fill = red_fill 
                    
                
        new_sheet.merge_cells('AJ11:AQ11')
        new_sheet.merge_cells('AR11:AY11')
        new_sheet.merge_cells('BC11:BJ11')
        new_sheet.merge_cells('BL11:BS11')       
        
        print("Datos escritos en la hoja con éxito.")
        
        # Save the workbook to the new path
        book.save(f"{new_path}/Summary_inspection_line_Trad_Report_{month}.xlsx")
        print("Cambios guardados exitosamente.") 
        
        # Close the workbook
        book.close()
    except FileNotFoundError:
        print("El archivo no fue encontrado en la ruta especificada.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")

